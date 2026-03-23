"""
RiKi Global — HPLC Column Quality Control Statistical Analysis
Data Analyst Portfolio Project 2

Applies pharmaceutical-standard statistical process control (SPC) to
HPLC column batch performance data. Detects OOS results, computes
capability indices (Cpk), and generates control charts.
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from scipy import stats
import os
import warnings
warnings.filterwarnings("ignore")

BASE = os.path.dirname(__file__)
REPORT_DIR = os.path.join(BASE, "report")
os.makedirs(REPORT_DIR, exist_ok=True)

np.random.seed(0)

# ── Simulate HPLC QC Dataset ─────────────────────────────────
brands = {"ACE": (0.98, 0.01), "Hichrom": (0.96, 0.015), "Cogent": (0.97, 0.012)}
column_types = ["C18", "C8", "Phenyl", "NH2", "CN"]
batch_count = 30
records = []

for brand, (mu_eff, sd_eff) in brands.items():
    for col_type in column_types:
        for batch in range(1, batch_count + 1):
            n_tests = np.random.randint(5, 12)
            # Plate count: USP spec ≥ 2000, target ~12000
            N_mean = np.random.normal(11500 * mu_eff, 800)
            # Tailing factor: USP spec 0.8–2.0, target ~1.1
            Tf_mean = np.random.normal(1.1, 0.1 * (1 + sd_eff * 5))
            # Resolution: spec ≥ 2.0, target ~4.5
            Rs_mean = np.random.normal(4.5 * mu_eff, 0.4)
            # Back pressure: spec ≤ 400 bar, target ~180 bar
            BP_mean = np.random.normal(185, 15 * (1 + sd_eff * 3))

            for i in range(n_tests):
                # Inject occasional failures
                fail_flag = np.random.random() < 0.04
                records.append({
                    "brand": brand,
                    "column_type": col_type,
                    "batch_id": f"{brand[:3]}-{col_type}-B{batch:02d}",
                    "test_id": f"T{i+1:02d}",
                    "plate_count_N": max(1000, round(np.random.normal(N_mean, 400) * (0.6 if fail_flag else 1))),
                    "tailing_factor": round(np.random.normal(Tf_mean, 0.08), 3),
                    "resolution_Rs": round(max(0.5, np.random.normal(Rs_mean, 0.3)), 2),
                    "back_pressure_bar": round(np.random.normal(BP_mean, 10), 1),
                    "analyst": np.random.choice(["Analyst_A", "Analyst_B", "Analyst_C"]),
                    "instrument": np.random.choice(["HPLC_01", "HPLC_02", "HPLC_03"]),
                })

df = pd.DataFrame(records)

# ── USP Spec Limits ───────────────────────────────────────────
SPECS = {
    "plate_count_N":     {"LSL": 2000,  "USL": None, "target": 12000},
    "tailing_factor":    {"LSL": 0.80,  "USL": 2.00, "target": 1.10},
    "resolution_Rs":     {"LSL": 2.00,  "USL": None, "target": 4.50},
    "back_pressure_bar": {"LSL": None,  "USL": 400,  "target": 185},
}

def oos_flag(row):
    flags = []
    for param, lims in SPECS.items():
        v = row[param]
        if lims["LSL"] and v < lims["LSL"]: flags.append(f"{param}<LSL")
        if lims["USL"] and v > lims["USL"]: flags.append(f"{param}>USL")
    return "; ".join(flags) if flags else "PASS"

df["oos_reason"] = df.apply(oos_flag, axis=1)
df["status"] = df["oos_reason"].apply(lambda x: "FAIL" if x != "PASS" else "PASS")

# Save dataset
df.to_csv(os.path.join(BASE, "data", "hplc_qc_data.csv"), index=False)
print(f"Generated {len(df):,} QC test records")
print(f"OOS Rate: {(df['status']=='FAIL').mean()*100:.1f}%")

# ── Cpk Calculation ───────────────────────────────────────────
def cpk(series, lsl=None, usl=None):
    mu, sigma = series.mean(), series.std()
    if sigma == 0: return np.nan
    cpu = (usl - mu) / (3 * sigma) if usl else np.inf
    cpl = (mu - lsl) / (3 * sigma) if lsl else np.inf
    return round(min(cpu, cpl), 3)

cpk_results = []
for brand in df["brand"].unique():
    for col_type in df["column_type"].unique():
        sub = df[(df["brand"] == brand) & (df["column_type"] == col_type)]
        cpk_results.append({
            "brand": brand, "column_type": col_type, "n": len(sub),
            "cpk_N":  cpk(sub["plate_count_N"], lsl=2000),
            "cpk_Tf": cpk(sub["tailing_factor"], lsl=0.8, usl=2.0),
            "cpk_Rs": cpk(sub["resolution_Rs"], lsl=2.0),
            "oos_rate_pct": round((sub["status"]=="FAIL").mean()*100, 1),
        })
cpk_df = pd.DataFrame(cpk_results)

# ── Figure 1: QC Dashboard ────────────────────────────────────
fig = plt.figure(figsize=(18, 12))
fig.suptitle("RiKi Global — HPLC Column QC Statistical Dashboard",
             fontsize=15, fontweight="bold", y=0.99)
gs = gridspec.GridSpec(2, 3, figure=fig, hspace=0.4, wspace=0.35)

COLORS = {"ACE": "#1a6b9a", "Hichrom": "#2ecc71", "Cogent": "#e74c3c"}

# A – Plate Count Distribution by Brand
ax1 = fig.add_subplot(gs[0, 0])
for brand, col in COLORS.items():
    vals = df[df["brand"]==brand]["plate_count_N"]
    ax1.hist(vals, bins=30, alpha=0.6, color=col, label=brand)
ax1.axvline(2000, color="red", linestyle="--", lw=2, label="LSL=2000")
ax1.set_title("Plate Count Distribution", fontweight="bold")
ax1.set_xlabel("Plate Count (N)"); ax1.set_ylabel("Frequency")
ax1.legend(fontsize=8)

# B – Tailing Factor Box Plot
ax2 = fig.add_subplot(gs[0, 1])
data_by_brand = [df[df["brand"]==b]["tailing_factor"].values for b in COLORS]
bp = ax2.boxplot(data_by_brand, labels=list(COLORS.keys()),
                 patch_artist=True, notch=True)
for patch, col in zip(bp["boxes"], COLORS.values()):
    patch.set_facecolor(col); patch.set_alpha(0.7)
ax2.axhline(0.8, color="red", ls="--", lw=1.5, label="LSL=0.8")
ax2.axhline(2.0, color="red", ls="--", lw=1.5, label="USL=2.0")
ax2.set_title("Tailing Factor — Brand Comparison", fontweight="bold")
ax2.set_ylabel("Tailing Factor (Tf)")
ax2.legend(fontsize=8)

# C – OOS Rate Heatmap
ax3 = fig.add_subplot(gs[0, 2])
pivot = cpk_df.pivot(index="brand", columns="column_type", values="oos_rate_pct")
im = ax3.imshow(pivot.values, cmap="RdYlGn_r", aspect="auto", vmin=0, vmax=10)
ax3.set_xticks(range(len(pivot.columns))); ax3.set_xticklabels(pivot.columns, rotation=30)
ax3.set_yticks(range(len(pivot.index))); ax3.set_yticklabels(pivot.index)
for i in range(pivot.shape[0]):
    for j in range(pivot.shape[1]):
        ax3.text(j, i, f"{pivot.iloc[i,j]:.1f}%", ha="center", va="center", fontsize=9)
plt.colorbar(im, ax=ax3, label="OOS Rate (%)")
ax3.set_title("OOS Rate Heatmap\n(Brand × Column Type)", fontweight="bold")

# D – SPC X-bar chart for ACE C18 plate count
ax4 = fig.add_subplot(gs[1, :2])
ace_c18 = df[(df["brand"]=="ACE") & (df["column_type"]=="C18")].groupby("batch_id")["plate_count_N"].mean().reset_index()
vals = ace_c18["plate_count_N"].values
grand_mean = vals.mean()
ucl = grand_mean + 3 * vals.std()
lcl = max(2000, grand_mean - 3 * vals.std())
ax4.plot(range(len(vals)), vals, marker="o", color="#1a6b9a", lw=1.5, ms=5)
ax4.axhline(grand_mean, color="green", lw=1.5, ls="-", label=f"X̄={grand_mean:.0f}")
ax4.axhline(ucl, color="red", lw=1.5, ls="--", label=f"UCL={ucl:.0f}")
ax4.axhline(lcl, color="red", lw=1.5, ls="--", label=f"LCL/LSL={lcl:.0f}")
ax4.axhline(2000, color="darkred", lw=2, ls=":", label="USP LSL=2000")
out_idx = [i for i, v in enumerate(vals) if v < lcl or v > ucl]
ax4.scatter(out_idx, vals[out_idx], color="red", zorder=5, s=80, label="OOC Points")
ax4.set_title("SPC X-bar Control Chart — ACE C18 Plate Count (Batch Means)", fontweight="bold")
ax4.set_xlabel("Batch"); ax4.set_ylabel("Plate Count (N)")
ax4.legend(fontsize=8, ncol=3)

# E – Cpk Bar Chart
ax5 = fig.add_subplot(gs[1, 2])
brands_u = cpk_df["brand"].unique()
x = np.arange(len(brands_u))
w = 0.25
for i, (param, col) in enumerate(zip(["cpk_N","cpk_Tf","cpk_Rs"], ["#1a6b9a","#e74c3c","#2ecc71"])):
    means = [cpk_df[cpk_df["brand"]==b][param].mean() for b in brands_u]
    ax5.bar(x + i*w, means, w, label=param.replace("cpk_","Cpk "), color=col, alpha=0.8)
ax5.axhline(1.33, color="black", ls="--", lw=1.5, label="Cpk=1.33 (GMP)")
ax5.set_xticks(x + w); ax5.set_xticklabels(brands_u)
ax5.set_title("Process Capability (Cpk)\nby Brand & Parameter", fontweight="bold")
ax5.set_ylabel("Cpk"); ax5.legend(fontsize=8)

plt.savefig(os.path.join(REPORT_DIR, "qc_dashboard.png"), dpi=150, bbox_inches="tight")
plt.close()
print("✓ Saved qc_dashboard.png")

# ── ANOVA: Is plate count significantly different across brands? ─
groups = [df[df["brand"]==b]["plate_count_N"].values for b in df["brand"].unique()]
f_stat, p_val = stats.f_oneway(*groups)
print(f"\nOne-way ANOVA — Plate Count across brands: F={f_stat:.2f}, p={p_val:.4f}")
print(f"  → {'Significant difference' if p_val < 0.05 else 'No significant difference'} (α=0.05)")

# ── Excel QC Report ───────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage

wb = Workbook()

def add_sheet(wb, df, name, header_color="1a6b9a"):
    ws = wb.create_sheet(name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    fill = PatternFill("solid", fgColor=header_color)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = \
            max(len(str(col[0].value or "")), max(len(str(c.value or "")) for c in col)) + 3
    return ws

del wb["Sheet"]

ws_chart = wb.create_sheet("QC Dashboard", 0)
ws_chart.sheet_view.showGridLines = False
img = XLImage(os.path.join(REPORT_DIR, "qc_dashboard.png"))
img.anchor = "A1"
ws_chart.add_image(img)

oos_df = df[df["status"]=="FAIL"][["brand","column_type","batch_id","test_id",
    "plate_count_N","tailing_factor","resolution_Rs","oos_reason"]].reset_index(drop=True)
add_sheet(wb, oos_df, "OOS Records", "C00000")
add_sheet(wb, cpk_df, "Cpk Summary", "1a6b9a")
add_sheet(wb, df.describe().reset_index(), "Descriptive Stats", "2e7d32")

wb.save(os.path.join(REPORT_DIR, "qc_report.xlsx"))
print("✓ Saved qc_report.xlsx")
